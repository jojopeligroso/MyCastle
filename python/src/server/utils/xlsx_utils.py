"""
XLSX file parsing utilities for Archon transform feature

Handles:
- XLSX file parsing with openpyxl
- Sheet enumeration and selection
- Data extraction to structured format
- Column type detection
"""

import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..config.logfire_config import get_logger

logger = get_logger(__name__)


class XLSXParseError(Exception):
    """Raised when XLSX file parsing fails"""

    pass


def parse_xlsx_file(file_path: str | Path) -> dict[str, Any]:
    """
    Parse an XLSX file and extract all sheets with metadata

    Args:
        file_path: Path to the XLSX file

    Returns:
        Dict containing sheets, metadata, and file info

    Raises:
        XLSXParseError: If file cannot be parsed
    """
    try:
        file_path = Path(file_path)
        if not file_path.exists():
            raise XLSXParseError(f"File not found: {file_path}")

        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)

        sheets = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Extract sheet metadata
            sheet_info = {
                "name": sheet_name,
                "row_count": sheet.max_row,
                "column_count": sheet.max_column,
                "has_data": sheet.max_row > 0,
            }

            # Try to detect if sheet name contains a date
            date_match = re.search(r"(\d{4}[-/]\d{2}[-/]\d{2})", sheet_name)
            if date_match:
                try:
                    sheet_info["date_from_name"] = datetime.strptime(
                        date_match.group(1).replace("/", "-"), "%Y-%m-%d"
                    ).isoformat()
                except ValueError:
                    pass

            sheets.append(sheet_info)

        workbook.close()

        # Sort sheets by date if available, otherwise keep original order
        sheets_with_dates = [s for s in sheets if "date_from_name" in s]
        sheets_without_dates = [s for s in sheets if "date_from_name" not in s]

        if sheets_with_dates:
            sheets_with_dates.sort(key=lambda x: x["date_from_name"], reverse=True)
            sorted_sheets = sheets_with_dates + sheets_without_dates
        else:
            sorted_sheets = sheets

        return {
            "file_name": file_path.name,
            "file_size": file_path.stat().st_size,
            "sheet_count": len(sheets),
            "sheets": sorted_sheets,
            "most_recent_sheet": sorted_sheets[0]["name"] if sorted_sheets else None,
        }

    except InvalidFileException as e:
        raise XLSXParseError(f"Invalid XLSX file: {e}") from e
    except Exception as e:
        logger.error(f"Failed to parse XLSX file: {e}", exc_info=True)
        raise XLSXParseError(f"Failed to parse XLSX file: {e}") from e


def extract_sheet_data(file_path: str | Path, sheet_name: str, max_rows: int = 1000) -> dict[str, Any]:
    """
    Extract data from a specific sheet in an XLSX file

    Args:
        file_path: Path to the XLSX file
        sheet_name: Name of the sheet to extract
        max_rows: Maximum number of rows to extract (including header)

    Returns:
        Dict containing columns, rows, and metadata

    Raises:
        XLSXParseError: If sheet cannot be extracted
    """
    try:
        file_path = Path(file_path)
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)

        if sheet_name not in workbook.sheetnames:
            raise XLSXParseError(f"Sheet '{sheet_name}' not found in workbook")

        sheet = workbook[sheet_name]

        # Extract all data
        data = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx >= max_rows:
                break
            # Convert None to empty string for consistency
            data.append([cell if cell is not None else "" for cell in row])

        workbook.close()

        if not data:
            raise XLSXParseError(f"Sheet '{sheet_name}' is empty")

        # First row is assumed to be headers
        headers = data[0] if data else []
        rows = data[1:] if len(data) > 1 else []

        # Clean headers - remove empty columns at the end
        while headers and not headers[-1]:
            headers.pop()
            rows = [row[: len(headers)] for row in rows]

        # Detect column types
        column_types = _detect_column_types(headers, rows)

        return {
            "sheet_name": sheet_name,
            "columns": headers,
            "column_types": column_types,
            "row_count": len(rows),
            "rows": rows[:100],  # Only return first 100 rows for preview
            "total_rows": len(rows),
            "has_more": len(rows) > 100,
        }

    except Exception as e:
        logger.error(f"Failed to extract sheet data: {e}", exc_info=True)
        raise XLSXParseError(f"Failed to extract sheet data: {e}") from e


def _detect_column_types(headers: list[str], rows: list[list[Any]]) -> dict[str, str]:
    """
    Detect the data type of each column based on content

    Args:
        headers: List of column headers
        rows: List of data rows

    Returns:
        Dict mapping column name to detected type
    """
    column_types = {}

    for col_idx, header in enumerate(headers):
        if not header:
            continue

        # Sample first 10 non-empty values
        sample_values = []
        for row in rows[:20]:  # Check first 20 rows
            if col_idx < len(row) and row[col_idx] != "":
                sample_values.append(row[col_idx])
                if len(sample_values) >= 10:
                    break

        if not sample_values:
            column_types[header] = "text"
            continue

        # Detect type based on samples
        detected_type = "text"

        # Check if all values are numbers
        if all(isinstance(v, (int, float)) for v in sample_values):
            detected_type = "number"
        # Check if all values are dates
        elif all(isinstance(v, datetime) for v in sample_values):
            detected_type = "date"
        # Check if values look like dates (string format)
        elif all(isinstance(v, str) and _looks_like_date(v) for v in sample_values):
            detected_type = "date"
        # Check if values are booleans
        elif all(isinstance(v, bool) for v in sample_values):
            detected_type = "boolean"

        column_types[header] = detected_type

    return column_types


def _looks_like_date(value: str) -> bool:
    """Check if a string value looks like a date"""
    date_patterns = [
        r"\d{4}[-/]\d{2}[-/]\d{2}",  # YYYY-MM-DD or YYYY/MM/DD
        r"\d{2}[-/]\d{2}[-/]\d{4}",  # DD-MM-YYYY or DD/MM/YYYY
        r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}",  # Flexible date format
    ]

    for pattern in date_patterns:
        if re.match(pattern, str(value).strip()):
            return True
    return False
